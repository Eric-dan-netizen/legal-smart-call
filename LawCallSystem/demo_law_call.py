import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import time
import random
import openpyxl
from openpyxl import Workbook

# ------------------ 模拟话术库 ------------------
TALK_SCRIPTS = {
    "高意向": "您好，我是XX律所的顾问，关于您咨询的案件，我们初步评估很有希望。建议您尽快来律所面谈，我可以帮您预约律师时间。",
    "一般意向": "您好，我是XX律所的顾问，关于您咨询的案件，我们还需要进一步了解情况。方便的话可以加个微信，我发资料给您参考。",
    "无意向": "您好，打扰了，如果您后续有法律需求，可以随时联系我们，祝您生活愉快。"
}

class DemoCallSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("律所智能外呼系统（演示版）")
        self.root.geometry("1100x700")
        self.root.resizable(True, True)

        self.customer_list = []
        self.call_records = []
        self.is_calling = False
        self.current_thread = None

        self.create_widgets()

    def create_widgets(self):
        top_frame = tk.Frame(self.root, padx=10, pady=10)
        top_frame.pack(fill=tk.X)

        btn_import = tk.Button(top_frame, text="📂 导入Excel客户名单", command=self.load_excel, width=18)
        btn_import.pack(side=tk.LEFT, padx=5)

        btn_start = tk.Button(top_frame, text="▶ 开始自动外呼", command=self.start_call, bg="#2ecc71", fg="white", width=18)
        btn_start.pack(side=tk.LEFT, padx=5)

        btn_stop = tk.Button(top_frame, text="⏸ 停止外呼", command=self.stop_call, bg="#e74c3c", fg="white", width=18)
        btn_stop.pack(side=tk.LEFT, padx=5)

        btn_export = tk.Button(top_frame, text="💾 导出意向客户", command=self.export_intent, width=18)
        btn_export.pack(side=tk.LEFT, padx=5)

        table_frame = tk.Frame(self.root, padx=10, pady=5)
        table_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("序号", "手机号", "意向等级", "通话状态", "通话摘要")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=140, anchor=tk.CENTER)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.status_var = tk.StringVar(value="就绪 | 未导入客户")
        status_bar = tk.Label(self.root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def load_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx *.xls")])
        if not file_path:
            return
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            self.customer_list = [str(cell.value) for cell in ws[ws.cell(1,1).column] if cell.value]
            self.call_records = []
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.status_var.set(f"已导入 {len(self.customer_list)} 个客户号码")
            messagebox.showinfo("导入成功", f"共导入 {len(self.customer_list)} 条客户号码")
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{str(e)}")

    def start_call(self):
        if not self.customer_list:
            messagebox.showwarning("提示", "请先导入客户名单")
            return
        if self.is_calling:
            messagebox.showinfo("提示", "外呼任务已在运行")
            return
        self.is_calling = True
        self.status_var.set("正在外呼...")
        self.current_thread = threading.Thread(target=self._call_worker, daemon=True)
        self.current_thread.start()

    def stop_call(self):
        self.is_calling = False
        self.status_var.set("外呼已停止")

    def _call_worker(self):
        for idx, phone in enumerate(self.customer_list):
            if not self.is_calling:
                break
            self.status_var.set(f"正在呼叫：{phone}")
            time.sleep(0.5)

            r = random.random()
            if r < 0.3:
                intent = "高意向"
            elif r < 0.7:
                intent = "一般意向"
            else:
                intent = "无意向"

            script = TALK_SCRIPTS.get(intent, "常规话术")
            call_summary = f"AI对话：{script[:40]}..." if len(script)>40 else script

            self.tree.insert("", "end", values=(idx+1, phone, intent, "已接通", call_summary))
            self.call_records.append({"手机号": phone, "意向等级": intent, "通话摘要": call_summary})
            self.root.update_idletasks()

        if self.is_calling:
            self.status_var.set("✅ 外呼任务已完成")
        self.is_calling = False

    def export_intent(self):
        if not self.call_records:
            messagebox.showwarning("提示", "没有通话记录，无法导出")
            return

        high_intent = [r for r in self.call_records if r["意向等级"] in ("高意向", "一般意向")]
        if not high_intent:
            messagebox.showinfo("提示", "没有筛选出意向客户")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel文件", "*.xlsx")])
        if save_path:
            wb = Workbook()
            ws = wb.active
            ws.title = "意向客户"
            ws.append(["手机号", "意向等级", "通话摘要"])
            for row in high_intent:
                ws.append([row["手机号"], row["意向等级"], row["通话摘要"]])
            wb.save(save_path)
            messagebox.showinfo("导出成功", f"已导出 {len(high_intent)} 条意向客户到：{save_path}")

if __name__ == "__main__":
    root = tk.Tk()
    app = DemoCallSystem(root)
    root.mainloop()